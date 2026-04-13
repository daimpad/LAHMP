"""
LAHMP Excel → JSON export script.
Reads three Excel workbooks and writes four JSON files to data/.

Usage:
    pip install openpyxl
    python export/convert.py
"""

import json
import re
import sys
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent.parent
DATA = BASE / "data"
DATA.mkdir(exist_ok=True)

# ---------------------------------------------------------------------------
# Helper parsers
# ---------------------------------------------------------------------------

def parse_int_list(val):
    """'1, 2, 3'  →  [1, 2, 3].  Returns None for blank/null cells."""
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() in ("none", "not applicable", "n/a", "-"):
        return None
    nums = re.findall(r"\d+", s)
    return [int(n) for n in nums] if nums else None


def parse_pcode_list(val):
    """'P01, P04, P08'  →  ['P01', 'P04', 'P08'].  Returns None for blank."""
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() in ("none", "not applicable", "n/a", "-"):
        return None
    codes = re.findall(r"P\d{2}", s)
    return codes if codes else None


def parse_efg_list(val):
    """'T7.1, T7.2 — Universal'  →  ['T7.1', 'T7.2'].  Returns None for blank."""
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() in ("none", "not applicable", "n/a", "-"):
        return None
    codes = re.findall(r"[A-Z]\d+\.\d+", s)
    return codes if codes else None


def parse_csv_str_list(val):
    """'A, B, C'  →  ['A', 'B', 'C'].  Returns None for blank."""
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() in ("none", "not applicable", "n/a", "-"):
        return None
    parts = [p.strip() for p in s.split(",")]
    result = [p for p in parts if p]
    return result if result else None


def parse_linkage_c(val):
    """'02 Mycorrhizal fungi, 06 Earthworms'  →  [2, 6]."""
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() in ("none", "not applicable", "n/a", "-"):
        return None
    # Leading zero-padded profile numbers, e.g. "02", "06", "F03"
    nums = re.findall(r"\b(\d{2})\b", s)
    return [int(n) for n in nums] if nums else None


def parse_prescreen_q(val):
    """'Q4 (description…)'  →  'Q4'.  'None' / blank  →  None."""
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower().startswith("none"):
        return None
    m = re.match(r"(Q\d)", s, re.IGNORECASE)
    return m.group(1).upper() if m else None


def clean_null(val, none_strings=("none", "not applicable", "n/a", "-", "")):
    """Return None if val is a recognised null-synonym, otherwise str(val).strip()."""
    if val is None:
        return None
    s = str(val).strip()
    return None if s.lower() in none_strings else s


def parse_protocol_level(val):
    """'Level 2 (field sampling…)'  →  2.  Returns lowest mentioned level."""
    if val is None:
        return None
    nums = re.findall(r"\bLevel\s*(\d)", str(val), re.IGNORECASE)
    return int(nums[0]) if nums else None


# ---------------------------------------------------------------------------
# EFG lookup (IUCN Global Ecosystem Typology — agricultural biome)
# ---------------------------------------------------------------------------

EFG_LOOKUP = {
    "T7.1": {"name": "Annual croplands", "biome": "Intensive land-use biome"},
    "T7.2": {"name": "Sown pastures and fields", "biome": "Intensive land-use biome"},
    "T7.3": {"name": "Plantations", "biome": "Intensive land-use biome"},
    "T7.4": {"name": "Urban and industrial ecosystems", "biome": "Intensive land-use biome"},
    "T7.5": {"name": "Derived semi-natural pastures and old fields", "biome": "Intensive land-use biome"},
    "F3.3": {"name": "Rice paddies", "biome": "Intensive land-use biome"},
    "F3.4": {"name": "Freshwater aquaculture systems", "biome": "Intensive land-use biome"},
    "FM1.3": {"name": "Coastal saltmarshes and reedbeds", "biome": "Transitional and ephemeral"},
    "T6.5": {"name": "Tropical shrublands and savannas", "biome": "Tropical-subtropical shrublands"},
    "T4.1": {"name": "Tropical savanna woodlands", "biome": "Tropical-subtropical grasslands"},
    "T4.2": {"name": "Pyric tussock savannas", "biome": "Tropical-subtropical grasslands"},
    "T2.1": {"name": "Boreal and montane forests", "biome": "Temperate-boreal forests"},
    "T2.2": {"name": "Temperate eucalypt forests", "biome": "Temperate-boreal forests"},
    "T2.4": {"name": "Warm temperate rainforests", "biome": "Temperate-boreal forests"},
    "M1.1": {"name": "Seagrass meadows", "biome": "Marine shelf"},
    "MFT1.2": {"name": "Intertidal forests and shrublands", "biome": "Transitional and ephemeral"},
}

IPCC_SOIL_TYPES = [
    "Mineral Soils — High Activity Clay (HAC)",
    "Mineral Soils — Low Activity Clay (LAC)",
    "Mineral Soils — Sandy Soils",
    "Mineral Soils — Volcanic Soils (Andosols)",
    "Mineral Soils — Spodosols/Podzols",
    "Organic Soils (Histosols/Peatlands)",
    "Wetland Soils (Gleysols/Fluvisols)",
    "Arid/Semi-arid Soils (Aridisols/Calcisols)",
    "Halomorphic Soils (Solonetz/Solonchak)",
    "Lithosols/Shallow Soils",
]

# Universal baseline abiotic indicator IDs (from Universal Baseline Package sheet)
UNIVERSAL_BASELINE_IDS = {"A01", "A02", "A03", "A04", "A05", "A06", "A08", "A09"}

# ---------------------------------------------------------------------------
# 1. data/practices.json
# ---------------------------------------------------------------------------

def export_practices():
    wb = openpyxl.load_workbook(BASE / "LAHMP_Practice_Matrix.xlsx")
    ws = wb["Practice Matrix"]

    practices = []
    for row in range(2, ws.max_row + 1):
        p_code = clean_null(ws.cell(row, 1).value)
        if not p_code:
            continue

        practice = {
            "p_code":                       p_code,
            "name":                         clean_null(ws.cell(row, 2).value),
            "theme":                        clean_null(ws.cell(row, 3).value),
            "rationale":                    clean_null(ws.cell(row, 4).value),
            "approach_origins":             clean_null(ws.cell(row, 5).value),
            "block4_pressures":             parse_int_list(ws.cell(row, 6).value),
            "block5_challenges":            parse_int_list(ws.cell(row, 7).value),
            "block6_services":              parse_int_list(ws.cell(row, 8).value),
            "hard_prerequisites":           clean_null(ws.cell(row, 9).value),
            "primary_applicability":        parse_csv_str_list(ws.cell(row, 10).value),
            "transformative_applicability": parse_csv_str_list(ws.cell(row, 11).value),
            "prescreen_question":           parse_prescreen_q(ws.cell(row, 12).value),
            "applicable_scale":             clean_null(ws.cell(row, 13).value),
            "relevant_efgs":                parse_efg_list(ws.cell(row, 14).value),
            "implementation_constraints":   clean_null(ws.cell(row, 15).value),
            "tape_mapping":                 clean_null(ws.cell(row, 16).value),
            "field_observation_checklist":  clean_null(ws.cell(row, 17).value),
            "evidence_correct":             clean_null(ws.cell(row, 18).value),
            "evidence_non_implementation":  clean_null(ws.cell(row, 19).value),
        }
        practices.append(practice)

    out = DATA / "practices.json"
    out.write_text(json.dumps(practices, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"  practices.json  ->  {len(practices)} practices")
    return practices


# ---------------------------------------------------------------------------
# 2. data/indicators.json
# ---------------------------------------------------------------------------

def export_indicators():
    wb = openpyxl.load_workbook(BASE / "LAHMP_Indicator_Linkage_Matrix_Populated.xlsx")
    ws = wb["Indicator Linkage Matrix"]

    indicators = []
    for row in range(2, ws.max_row + 1):
        raw_num = ws.cell(row, 1).value
        if raw_num is None:
            continue

        # Profile number: '01' or '06' etc. → integer
        profile_number_str = str(raw_num).strip()
        try:
            profile_number = int(profile_number_str)
        except ValueError:
            # Skip rows that don't have an integer profile number
            continue

        validation_status = clean_null(ws.cell(row, 30).value) or "DRAFT - Unvalidated"

        # A profile is considered "populated" if it has protocol content
        level1_protocol = clean_null(ws.cell(row, 21).value)
        b2_practices   = parse_pcode_list(ws.cell(row, 19).value)
        block4         = parse_int_list(ws.cell(row, 11).value)
        populated = bool(level1_protocol or b2_practices or block4)

        indicator = {
            "profile_number":                profile_number,
            "profile_name":                  clean_null(ws.cell(row, 2).value),
            "category":                      clean_null(ws.cell(row, 3).value),
            "tier":                          clean_null(ws.cell(row, 4).value),
            "conditionality_criteria":       clean_null(ws.cell(row, 5).value,
                                                 none_strings=("none", "not applicable", "n/a",
                                                               "-", "", "not applicable")),
            "hard_prerequisite":             clean_null(ws.cell(row, 6).value),
            "primary_monitoring_role":       clean_null(ws.cell(row, 8).value),
            "monitoring_stage":              clean_null(ws.cell(row, 9).value),
            "response_timescale":            clean_null(ws.cell(row, 10).value),
            "block4_pressures":              parse_int_list(ws.cell(row, 11).value),
            "block5_challenges":             parse_int_list(ws.cell(row, 12).value),
            "block6_services":               parse_int_list(ws.cell(row, 13).value),
            "relevant_efgs":                 parse_efg_list(ws.cell(row, 14).value),
            "relevant_ipcc_land_use":        parse_csv_str_list(ws.cell(row, 15).value),
            "soil_type_associations":        parse_csv_str_list(ws.cell(row, 16).value),
            "crop_livestock_associations":   parse_csv_str_list(ws.cell(row, 17).value),
            "b1_practices_that_benefit":     parse_pcode_list(ws.cell(row, 18).value),
            "b2_practices_primarily_verified": b2_practices,
            "linkage_c_connected_groups":    parse_linkage_c(ws.cell(row, 20).value),
            "level1_protocol_name":          level1_protocol,
            "level2_protocol_name":          clean_null(ws.cell(row, 22).value),
            "level3_protocol_name":          clean_null(ws.cell(row, 23).value),
            "level1_output_metric":          clean_null(ws.cell(row, 24).value),
            "level2_output_metric":          clean_null(ws.cell(row, 25).value),
            "level3_output_metric":          clean_null(ws.cell(row, 26).value),
            "primary_reference":             clean_null(ws.cell(row, 27).value),
            "monitoring_task_type":          clean_null(ws.cell(row, 28).value),
            "b2_expected_direction_of_change": clean_null(ws.cell(row, 29).value),
            "validation_status":             validation_status,
            "populated":                     populated,
        }
        indicators.append(indicator)

    out = DATA / "indicators.json"
    out.write_text(json.dumps(indicators, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"  indicators.json ->  {len(indicators)} profiles  "
          f"({sum(1 for i in indicators if i['populated'])} populated)")
    return indicators


# ---------------------------------------------------------------------------
# 3. data/abiotic.json
# ---------------------------------------------------------------------------

def export_abiotic():
    wb = openpyxl.load_workbook(BASE / "LAHMP_Abiotic_Reference_Table.xlsx")
    ws = wb["Abiotic Reference Table"]

    abiotic = []
    for row in range(2, ws.max_row + 1):
        ind_id = clean_null(ws.cell(row, 1).value)
        if not ind_id:
            continue

        abiotic.append({
            "indicator_id":             ind_id,
            "indicator_name":           clean_null(ws.cell(row, 2).value),
            "domain":                   clean_null(ws.cell(row, 3).value),
            "what_it_measures":         clean_null(ws.cell(row, 4).value),
            "block4_pressures":         parse_int_list(ws.cell(row, 8).value),
            "block5_challenges":        parse_int_list(ws.cell(row, 9).value),
            "block6_services":          None,          # not in Excel
            "linked_practices":         parse_pcode_list(ws.cell(row, 10).value),
            "hard_prerequisite_land_use": clean_null(ws.cell(row, 14).value),
            "universal_baseline":       ind_id in UNIVERSAL_BASELINE_IDS,
            "protocol_name":            clean_null(ws.cell(row, 13).value),
            "monitoring_frequency":     clean_null(ws.cell(row, 12).value),
            "equipment_required":       [],            # not in Excel
            "protocol_level":           parse_protocol_level(ws.cell(row, 13).value),
            "interpretation_notes":     clean_null(ws.cell(row, 15).value),
            "response_timescale":       clean_null(ws.cell(row, 11).value),
            "validation_status":        clean_null(ws.cell(row, 16).value),
        })

    out = DATA / "abiotic.json"
    out.write_text(json.dumps(abiotic, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"  abiotic.json    ->  {len(abiotic)} indicators  "
          f"({sum(1 for a in abiotic if a['universal_baseline'])} universal baseline)")
    return abiotic


# ---------------------------------------------------------------------------
# 4. data/reference.json
# ---------------------------------------------------------------------------

def _read_ref_list(ws, id_col, name_col, group_col, start_row=3):
    """Read a numbered list from a Reference Lists sheet."""
    items = []
    for row in range(start_row, ws.max_row + 1):
        id_val = ws.cell(row, id_col).value
        name   = ws.cell(row, name_col).value
        if id_val is None or name is None:
            continue
        try:
            item_id = int(str(id_val).strip())
        except ValueError:
            continue
        group = ws.cell(row, group_col).value
        items.append({
            "id":      item_id,
            "name":    str(name).strip(),
            "group":   str(group).strip() if group else None,
            "tooltip": "",
        })
    return items


def _derive_pressure_to_challenge(practices):
    """
    Derive pressure→challenge mapping from the Practice Matrix.
    For each practice, every (pressure, challenge) pair maps "high" confidence.
    Union logic: if the same (pressure, challenge) appears in multiple practices,
    keep the existing confidence (all are "high" here).
    """
    mapping = {}  # pressure_id → { challenge_id → confidence }
    for p in practices:
        pressures   = p.get("block4_pressures") or []
        challenges  = p.get("block5_challenges") or []
        for pid in pressures:
            if pid not in mapping:
                mapping[pid] = {}
            for cid in challenges:
                # high confidence for direct co-occurrence in a practice
                if cid not in mapping[pid]:
                    mapping[pid][cid] = "high"
    # Serialise to the expected format
    return {
        str(pid): [
            {"challenge_id": cid, "confidence": conf}
            for cid, conf in sorted(chals.items())
        ]
        for pid, chals in sorted(mapping.items())
    }


def _derive_challenge_to_service(practices):
    """Derive challenge→service mapping from the Practice Matrix."""
    mapping = {}
    for p in practices:
        challenges = p.get("block5_challenges") or []
        services   = p.get("block6_services") or []
        for cid in challenges:
            if cid not in mapping:
                mapping[cid] = {}
            for sid in services:
                if sid not in mapping[cid]:
                    mapping[cid][sid] = "high"
    return {
        str(cid): [
            {"service_id": sid, "confidence": conf}
            for sid, conf in sorted(svcs.items())
        ]
        for cid, svcs in sorted(mapping.items())
    }


IPCC_LAND_USE_CATEGORIES = [
    "Intensive Annual Cropland",
    "Extensive Annual Cropland",
    "Permanent Cropland",
    "Permanent Crops",
    "Irrigated Cropland",
    "Grassland and Pasture",
    "Grassland",
    "Degraded Grassland",
    "Mixed Crop-Livestock System",
    "Agroforestry",
    "Plantation Forestry",
    "Forest Land",
    "Shrubland",
    "Wetlands",
    "Other Land",
]


def _collect_ipcc_categories(practices):
    """Return the canonical IPCC land use category list."""
    return IPCC_LAND_USE_CATEGORIES


def _collect_efg_options(practices, indicators):
    """Collect all unique EFG codes referenced in practices and indicators."""
    seen = set()
    options = []
    for code_list in [
        (p.get("relevant_efgs") or []) for p in practices
    ] + [
        (i.get("relevant_efgs") or []) for i in indicators
    ]:
        for code in code_list:
            code = code.strip()
            if code and code not in seen:
                seen.add(code)
                info = EFG_LOOKUP.get(code, {"name": code, "biome": "Unknown"})
                options.append({"code": code, "name": info["name"], "biome": info["biome"]})
    options.sort(key=lambda x: x["code"])
    return options


def export_reference(practices, indicators):
    # Use the Practice Matrix Reference Lists as the canonical source
    wb_pm  = openpyxl.load_workbook(BASE / "LAHMP_Practice_Matrix.xlsx")
    ws_ref = wb_pm["Reference Lists"]

    # Use the Indicator Matrix Reference Lists for challenges (has the eutrophication addition)
    wb_ilm  = openpyxl.load_workbook(BASE / "LAHMP_Indicator_Linkage_Matrix_Populated.xlsx")
    ws_ref2 = wb_ilm["Reference Lists"]

    # Block 4 pressures: cols 1-3 (id, name, group), data rows 3+
    block4_pressures = _read_ref_list(ws_ref, id_col=1, name_col=2, group_col=3)

    # Block 5 challenges: use Indicator Matrix version (35 challenges incl. eutrophication)
    # In the ILM Reference Lists: col5=id, col6=name, col7=group
    block5_challenges = _read_ref_list(ws_ref2, id_col=5, name_col=6, group_col=7)

    # Block 6 services: cols 9-11 (id, name, group)
    block6_services = _read_ref_list(ws_ref, id_col=9, name_col=10, group_col=11)

    # Derive mappings from practice data
    pressure_to_challenge  = _derive_pressure_to_challenge(practices)
    challenge_to_service   = _derive_challenge_to_service(practices)

    ipcc_land_use_categories = _collect_ipcc_categories(practices)
    efg_options              = _collect_efg_options(practices, indicators)

    reference = {
        "block4_pressures":            block4_pressures,
        "block5_challenges":           block5_challenges,
        "block6_services":             block6_services,
        "ipcc_land_use_categories":    ipcc_land_use_categories,
        "ipcc_soil_types":             IPCC_SOIL_TYPES,
        "efg_options":                 efg_options,
        "pressure_to_challenge_mapping":  pressure_to_challenge,
        "challenge_to_service_mapping":   challenge_to_service,
    }

    out = DATA / "reference.json"
    out.write_text(json.dumps(reference, indent=2, ensure_ascii=False), encoding="utf-8")
    print(
        f"  reference.json  ->  "
        f"{len(block4_pressures)} pressures, "
        f"{len(block5_challenges)} challenges, "
        f"{len(block6_services)} services, "
        f"{len(efg_options)} EFG options"
    )
    return reference


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    print("LAHMP Excel -> JSON export")
    print(f"  Source: {BASE}")
    print(f"  Output: {DATA}")
    print()

    practices  = export_practices()
    indicators = export_indicators()
    abiotic    = export_abiotic()
    reference  = export_reference(practices, indicators)

    print()
    print("Done. All four JSON files written to data/.")

